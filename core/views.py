from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db import IntegrityError
from django.db.models import Max, Q, Count, Avg, Sum
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from .models import Usuario, Inventario, Asistencia, RegistroFalla, RegistroLlamada, Pedido, DetallePedido, Auditoria, ImagenCarrusel, Evento, Noticia, ManualInterno, Contacto
from .forms import CrearUsuarioForm, RegistroAsistenciaForm, CambiarPasswordForm, EditarAsistenciaForm, EditarUsuarioForm, RegistroFallaForm, RegistroLlamadaForm, CrearInventarioForm, EditarInventarioForm, CrearPedidoForm, EditarPrecioProductoForm, CambiarStockForm, ImagenCarruselForm, EventoForm, NoticiaForm, ContactoForm, ManualInternoForm
from django.http import HttpResponse
import json
import os
import logging
import traceback

logger = logging.getLogger(__name__)


def login_view(request):
    """Vista para el login - siempre accesible"""
    # Si el usuario ya está autenticado, redirigir según su rol y estado
    if request.user.is_authenticated:
        # Todos los usuarios deben cambiar su contraseña en el primer login
        if request.user.cambio_password_requerido:
            return redirect('cambiar_password')
        # Si es admin, ir al panel
        if request.user._es_administrador or request.user.is_superuser:
            return redirect('panel')
        # Si es usuario colaborador, redirigir a dashboard colaborador
        if request.user.es_colaborador:
            return redirect('colaborador_dashboard')
        # Fallback al panel
        return redirect('panel')
    
    if request.method == 'POST':
        account_type = request.POST.get('account_type')
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # Validar que se haya seleccionado un tipo de cuenta
        if not account_type:
            messages.error(request, 'Por favor selecciona un tipo de cuenta')
            return render(request, 'core/login.html')
        
        # Validar que se hayan proporcionado credenciales
        if not username or not password:
            messages.error(request, 'Por favor completa todos los campos')
            return render(request, 'core/login.html', {'account_type': account_type})
        
        # Autenticar usuario
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Verificar si el usuario está activo
            if not user.activo:
                messages.error(request, 'Tu cuenta está inactiva. Contacta al administrador para más información.')
                return render(request, 'core/login.html', {'account_type': account_type})
            
            # Validar que el tipo de cuenta seleccionado coincida con los roles del usuario
            is_gerencia = account_type == 'gerencia'
            is_colaborador = account_type == 'colaborador'
            
            # Verificar permisos según el tipo de cuenta seleccionado
            puede_acceder = False
            if is_gerencia and (user._es_administrador or user.is_superuser):
                puede_acceder = True
            elif is_colaborador and user.es_colaborador:
                puede_acceder = True
            # Si el usuario tiene ambos roles, puede acceder con cualquier tipo
            elif (user._es_administrador or user.is_superuser) and user.es_colaborador:
                puede_acceder = True
            
            if puede_acceder:
                login(request, user)
                
                # Todos los usuarios deben cambiar su contraseña en el primer login
                if user.cambio_password_requerido:
                    return redirect('cambiar_password')
                
                # Verificar si hay un parámetro 'next' en la URL
                next_url = request.GET.get('next', None)
                if next_url:
                    return redirect(next_url)
                
                # Redirigir según el tipo de cuenta seleccionado
                if is_gerencia:
                    # Si seleccionó Gerencia, ir al panel de administrador
                    return redirect('panel')
                elif is_colaborador:
                    # Si seleccionó Colaborador, ir al dashboard de colaborador
                    return redirect('colaborador_dashboard')
                
                # Fallback: Si tiene solo permisos de admin, ir al panel
                if (user._es_administrador or user.is_superuser) and not user.es_colaborador:
                    return redirect('panel')
                # Si tiene solo permisos de colaborador, ir al dashboard colaborador
                elif user.es_colaborador and not (user._es_administrador or user.is_superuser):
                    return redirect('colaborador_dashboard')
                
                # Fallback al panel si tiene ambos roles y no se seleccionó tipo específico
                return redirect('panel')
            else:
                # El tipo de cuenta no coincide con los roles del usuario
                messages.error(request, 'Usuario o Contraseña Incorrecto')
                return render(request, 'core/login.html', {'account_type': account_type})
        else:
            messages.error(request, 'Usuario o Contraseña Incorrecto')
            return render(request, 'core/login.html', {'account_type': account_type})
    
    return render(request, 'core/login.html')


def logout_view(request):
    """Vista para cerrar sesión - accesible sin autenticación para limpiar sesiones"""
    if request.user.is_authenticated:
        logout(request)
        messages.success(request, 'Sesión cerrada correctamente')
    
    # Siempre redirigir al index después de cerrar sesión
    return redirect('index')


def solicitar_restablecimiento_password(request):
    """Vista para solicitar restablecimiento de contraseña"""
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        
        if not username:
            messages.error(request, 'Por favor ingresa tu nombre de usuario o RUT')
            return render(request, 'core/solicitar_restablecimiento.html')
        
        # Buscar usuario por username o RUT
        try:
            usuario = Usuario.objects.get(Q(username=username) | Q(rut=username))
            
            # Verificar que el usuario esté activo
            if not usuario.activo:
                messages.error(request, 'Esta cuenta está desactivada. Contacta al administrador.')
                return render(request, 'core/solicitar_restablecimiento.html')
            
            # Registrar la solicitud en auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=usuario,
                accion='password_reset_request',
                modulo='sesion',
                descripcion=f'Usuario "{usuario.get_nombre_completo()}" (username: {usuario.username}) solicita restablecimiento de contraseña',
                objeto_afectado=f"{usuario.nombre} {usuario.apellido}" if usuario.nombre and usuario.apellido else usuario.username,
                detalles={
                    'username': usuario.username,
                    'rut': usuario.rut,
                    'correo': usuario.correo_institucional,
                }
            )
            
            messages.success(request, 'Tu solicitud de restablecimiento de contraseña ha sido registrada. Un administrador revisará tu solicitud pronto.')
            return redirect('login')
        except Usuario.DoesNotExist:
            # Por seguridad, mostrar el mismo mensaje aunque el usuario no exista
            messages.success(request, 'Si el usuario existe, se ha registrado la solicitud de restablecimiento de contraseña.')
            return redirect('login')
        except Usuario.MultipleObjectsReturned:
            # Esto no debería pasar, pero por si acaso
            messages.error(request, 'Error: múltiples usuarios encontrados. Contacta al administrador.')
            return render(request, 'core/solicitar_restablecimiento.html')
    
    return render(request, 'core/solicitar_restablecimiento.html')


@login_required
def restablecer_password_admin(request, usuario_id):
    """Vista para que el administrador restablezca la contraseña de un usuario"""
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    usuario = get_object_or_404(Usuario, id=usuario_id)
    
    if request.method == 'POST':
        # Establecer la contraseña por defecto
        usuario.set_password('popup')
        usuario.cambio_password_requerido = True
        usuario.save()
        
        # Registrar en auditoría
        from .utils import registrar_auditoria
        registrar_auditoria(
            usuario=request.user,
            accion='password_reset_admin',
            modulo='usuarios',
            descripcion=f'Administrador "{request.user.get_nombre_completo()}" restableció la contraseña de "{usuario.get_nombre_completo()}" (username: {usuario.username})',
            objeto_afectado=f"{usuario.nombre} {usuario.apellido}" if usuario.nombre and usuario.apellido else usuario.username,
            detalles={
                'usuario_afectado': usuario.username,
                'nueva_password': 'popup (por defecto)',
                'cambio_password_requerido': True,
            }
        )
        
        messages.success(request, f'Contraseña de "{usuario.get_nombre_completo()}" restablecida exitosamente. La contraseña por defecto es "popup" y el usuario deberá cambiarla en su próximo inicio de sesión.')
        return redirect('editar_usuario', usuario_id=usuario_id)
    
    return render(request, 'core/restablecer_password_confirm.html', {'usuario': usuario})


@login_required
def cambiar_password_view(request):
    """Vista para cambiar la contraseña cuando es requerido"""
    # Si el usuario no necesita cambiar su contraseña, redirigir al panel
    if not request.user.cambio_password_requerido:
        return redirect('panel')
    
    # Para usuarios colaboradores, permitir cambiar contraseña
    # Para administradores, también permitir si se restableció su contraseña
    if request.user.es_colaborador or (request.user._es_administrador and request.user.cambio_password_requerido):
        pass  # Continuar con el proceso
    else:
        return redirect('panel')
    
    form = CambiarPasswordForm(user=request.user)
    
    if request.method == 'POST':
        form = CambiarPasswordForm(user=request.user, data=request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Contraseña cambiada exitosamente. Por favor, inicie sesión nuevamente.')
            from django.contrib.auth import logout
            logout(request)
            return redirect('index')
    
    return render(request, 'core/cambiar_password.html', {'form': form})


@login_required
def panel_view(request):
    """Vista principal del panel de administrador"""
    # Todos los usuarios deben cambiar su contraseña si está requerido
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    
    # Solo administradores pueden acceder al panel
    if not (request.user._es_administrador or request.user.is_superuser):
        logout(request)
        return redirect('registro_asistencia')
    
    # Estadísticas generales
    stats = {
        'total_usuarios': Usuario.objects.count(),
    }
    
    return render(request, 'core/panel.html', {'stats': stats})


@login_required
def usuarios_view(request):
    """Vista para gestionar usuarios"""
    # Todos los usuarios deben cambiar su contraseña si está requerido
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    usuarios = Usuario.objects.all().order_by('-fecha_creacion')
    return render(request, 'core/usuarios.html', {'usuarios': usuarios})


@login_required
def crear_usuario_view(request):
    """Vista para crear usuarios desde el panel de administrador"""
    # Todos los usuarios deben cambiar su contraseña si está requerido
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    form = CrearUsuarioForm()
    
    if request.method == 'POST':
        form = CrearUsuarioForm(request.POST)
        if form.is_valid():
            try:
                usuario = form.save()
                messages.success(request, f'Usuario {usuario.get_nombre_completo()} creado exitosamente')
                return redirect('usuarios')
            except IntegrityError as e:
                messages.error(request, 'Error al crear el usuario. El RUT o correo ya existe.')
            except Exception as e:
                messages.error(request, f'Error al crear el usuario: {str(e)}')
    
    return render(request, 'core/crear_usuario.html', {'form': form})


@login_required
def editar_usuario_view(request, usuario_id):
    """Vista para editar un usuario desde el panel de administrador"""
    # Todos los usuarios deben cambiar su contraseña si está requerido
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    usuario = get_object_or_404(Usuario, id=usuario_id)
    
    # Verificar si el usuario intenta cambiar su propio estado y es administrador
    es_propia_cuenta_admin = usuario.id == request.user.id and (usuario._es_administrador or usuario.is_superuser)
    
    # Verificar si hay solicitudes pendientes de restablecimiento de contraseña para este usuario
    solicitudes_pendientes = Auditoria.objects.filter(
        accion='password_reset_request',
        objeto_afectado=f"{usuario.nombre} {usuario.apellido}" if usuario.nombre and usuario.apellido else usuario.username
    ).order_by('-fecha_hora')
    
    # Considerar una solicitud como "pendiente" si fue hecha en los últimos 30 días
    from datetime import timedelta
    fecha_limite = timezone.now() - timedelta(days=30)
    solicitudes_recientes = solicitudes_pendientes.filter(fecha_hora__gte=fecha_limite)
    
    # Verificar si ya se restableció la contraseña después de la última solicitud
    ultimo_restablecimiento = Auditoria.objects.filter(
        accion='password_reset_admin',
        objeto_afectado=f"{usuario.nombre} {usuario.apellido}" if usuario.nombre and usuario.apellido else usuario.username
    ).order_by('-fecha_hora').first()
    
    tiene_solicitud_pendiente = False
    if solicitudes_recientes.exists():
        if ultimo_restablecimiento:
            # Si hay un restablecimiento más reciente que la última solicitud, no mostrar como pendiente
            ultima_solicitud = solicitudes_recientes.first()
            if ultimo_restablecimiento.fecha_hora < ultima_solicitud.fecha_hora:
                tiene_solicitud_pendiente = True
        else:
            # Si no hay restablecimientos, mostrar como pendiente
            tiene_solicitud_pendiente = True
    
    if request.method == 'POST':
        form = EditarUsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            try:
                # Guardar datos anteriores para auditoría
                datos_anteriores = {
                    'nombre': usuario.nombre,
                    'apellido': usuario.apellido,
                    'rut': usuario.rut,
                    'correo_institucional': usuario.correo_institucional,
                    'es_administrador': usuario._es_administrador if hasattr(usuario, '_es_administrador') else (usuario.rol == 'admin' if usuario.rol else False),
                    'es_colaborador': usuario.es_colaborador if hasattr(usuario, 'es_colaborador') else (usuario.rol == 'usuario' if usuario.rol else True),
                    'activo': usuario.activo,
                }
                
                # Si es su propia cuenta de administrador, prevenir cambio de estado activo
                if es_propia_cuenta_admin:
                    # Restaurar el estado activo a True antes de guardar
                    usuario = form.save(commit=False)
                    usuario.activo = True
                    usuario.save()
                    messages.info(request, 'No puedes desactivar tu propia cuenta de administrador. El estado activo se mantuvo en True.')
                else:
                    usuario = form.save()
                
                # Refrescar desde la base de datos para obtener los valores actualizados
                usuario.refresh_from_db()
                
                # Registrar auditoría
                from .utils import registrar_auditoria
                detalles = {
                    'datos_anteriores': datos_anteriores,
                    'datos_nuevos': {
                        'nombre': usuario.nombre,
                        'apellido': usuario.apellido,
                        'rut': usuario.rut,
                        'correo_institucional': usuario.correo_institucional,
                        'es_administrador': usuario._es_administrador if hasattr(usuario, '_es_administrador') else False,
                        'es_colaborador': usuario.es_colaborador if hasattr(usuario, 'es_colaborador') else False,
                        'activo': usuario.activo,
                    }
                }
                registrar_auditoria(
                    usuario=request.user,
                    accion='usuario_edit',
                    modulo='usuarios',
                    descripcion=f'Usuario "{usuario.get_nombre_completo()}" (username: {usuario.username}) modificado',
                    objeto_afectado=f"{usuario.nombre} {usuario.apellido}",
                    detalles=detalles
                )
                messages.success(request, f'Usuario {usuario.get_nombre_completo()} actualizado exitosamente')
                return redirect('usuarios')
            except IntegrityError as e:
                messages.error(request, 'Error al actualizar el usuario. El RUT o correo ya existe.')
            except Exception as e:
                messages.error(request, f'Error al actualizar el usuario: {str(e)}')
    else:
        form = EditarUsuarioForm(instance=usuario)
        # Si es su propia cuenta de administrador, deshabilitar el campo activo
        if es_propia_cuenta_admin:
            form.fields['activo'].widget.attrs['disabled'] = True
            form.fields['activo'].widget.attrs['readonly'] = True
    
    return render(request, 'core/editar_usuario.html', {
        'form': form, 
        'usuario': usuario,
        'es_propia_cuenta_admin': es_propia_cuenta_admin,
        'tiene_solicitud_pendiente': tiene_solicitud_pendiente,
        'solicitudes_recientes': solicitudes_recientes
    })


@login_required
def eliminar_usuario_view(request, usuario_id):
    """Vista para eliminar un usuario (solo si no tiene registros relacionados)"""
    # Todos los usuarios deben cambiar su contraseña si está requerido
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    usuario = get_object_or_404(Usuario, id=usuario_id)
    
    # Verificar si el usuario intenta eliminar/desactivar su propia cuenta y es administrador
    if usuario.id == request.user.id and (usuario._es_administrador or usuario.is_superuser):
        messages.error(request, 'No puedes eliminar o desactivar tu propia cuenta de administrador.')
        return redirect('usuarios')
    
    # Verificar si el usuario tiene registros relacionados
    tiene_asistencias = usuario.asistencias.exists()
    tiene_fallas = usuario.fallas_registradas.exists()
    tiene_llamadas = usuario.llamadas_registradas.exists()
    tiene_pedidos = usuario.pedidos_creados.exists()
    tiene_auditorias = usuario.auditorias.exists()
    
    tiene_registros = tiene_asistencias or tiene_fallas or tiene_llamadas or tiene_pedidos or tiene_auditorias
    
    if request.method == 'POST':
        if tiene_registros:
            # Solo desactivar si tiene registros
            usuario.activo = False
            usuario.save()
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='usuario_deactivate',
                modulo='usuarios',
                descripcion=f'Usuario "{usuario.get_nombre_completo()}" desactivado (tiene registros relacionados)',
                objeto_afectado=f"{usuario.nombre} {usuario.apellido}",
                detalles={
                    'tiene_asistencias': tiene_asistencias,
                    'tiene_fallas': tiene_fallas,
                    'tiene_llamadas': tiene_llamadas,
                    'tiene_pedidos': tiene_pedidos,
                    'tiene_auditorias': tiene_auditorias,
                }
            )
            messages.warning(request, f'El usuario {usuario.get_nombre_completo()} tiene registros relacionados. Se ha desactivado en lugar de eliminar.')
            return redirect('usuarios')
        else:
            # Eliminar si no tiene registros
            nombre_completo = usuario.get_nombre_completo()
            usuario_nombre = f"{usuario.nombre} {usuario.apellido}"
            # Registrar auditoría antes de eliminar
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='usuario_delete',
                modulo='usuarios',
                descripcion=f'Usuario "{nombre_completo}" eliminado',
                objeto_afectado=usuario_nombre,
                detalles={
                    'username': usuario.username,
                    'rut': usuario.rut,
                    'correo_institucional': usuario.correo_institucional,
                }
            )
            usuario.delete()
            messages.success(request, f'Usuario {nombre_completo} eliminado exitosamente.')
            return redirect('usuarios')
    
    # Preparar información sobre los registros
    registros_info = []
    if tiene_asistencias:
        count = usuario.asistencias.count()
        registros_info.append(f'{count} registro(s) de asistencia')
    if tiene_fallas:
        count = usuario.fallas_registradas.count()
        registros_info.append(f'{count} registro(s) de fallas')
    if tiene_llamadas:
        count = usuario.llamadas_registradas.count()
        registros_info.append(f'{count} registro(s) de llamadas')
    if tiene_pedidos:
        count = usuario.pedidos_creados.count()
        registros_info.append(f'{count} pedido(s)')
    if tiene_auditorias:
        count = usuario.auditorias.count()
        registros_info.append(f'{count} registro(s) de auditoría')
    
    return render(request, 'core/eliminar_usuario.html', {
        'usuario': usuario,
        'tiene_registros': tiene_registros,
        'registros_info': registros_info
    })


@login_required
def inventario_view(request):
    """Vista para gestionar inventario"""
    # Todos los usuarios deben cambiar su contraseña si está requerido
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    # Organizar productos por categoría
    bodega = Inventario.objects.filter(categoria='bodega').order_by('nombre')
    meson = Inventario.objects.filter(categoria='meson').order_by('nombre')
    limpieza = Inventario.objects.filter(categoria='limpieza').order_by('nombre')
    
    context = {
        'bodega': bodega,
        'meson': meson,
        'limpieza': limpieza,
    }
    
    return render(request, 'core/inventario.html', context)


@login_required
def crear_inventario_view(request):
    """Vista para crear productos de inventario"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    form = CrearInventarioForm()
    
    if request.method == 'POST':
        form = CrearInventarioForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                producto = form.save()
                messages.success(request, f'Producto "{producto.nombre}" creado exitosamente.')
                return redirect('inventario')
            except Exception as e:
                messages.error(request, f'Error al crear el producto: {str(e)}')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    
    return render(request, 'core/crear_inventario.html', {'form': form})


@login_required
def editar_inventario_view(request, producto_id):
    """Vista para editar un producto de inventario"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    try:
        producto = Inventario.objects.get(id=producto_id)
    except Inventario.DoesNotExist:
        messages.error(request, "El producto no existe.")
        return redirect('inventario')
    
    if request.method == 'POST':
        form = EditarInventarioForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            # Guardar datos anteriores para auditoría
            datos_anteriores = {
                'nombre': producto.nombre,
                'categoria': producto.categoria,
                'cantidad': producto.cantidad,
            }
            form.save()
            producto.refresh_from_db()
            # Registrar auditoría
            from .utils import registrar_auditoria
            detalles = {
                'datos_anteriores': datos_anteriores,
                'datos_nuevos': {
                    'nombre': producto.nombre,
                    'categoria': producto.categoria,
                    'cantidad': producto.cantidad,
                }
            }
            registrar_auditoria(
                usuario=request.user,
                accion='inventario_edit',
                modulo='inventario',
                descripcion=f'Producto "{producto.nombre}" modificado',
                objeto_afectado=producto.nombre,
                detalles=detalles
            )
            messages.success(request, f'Producto "{producto.nombre}" actualizado correctamente.')
            return redirect('inventario')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = EditarInventarioForm(instance=producto)
    
    return render(request, 'core/editar_inventario.html', {
        'form': form,
        'producto': producto
    })


@login_required
def eliminar_inventario_view(request, producto_id):
    """Vista para eliminar un producto de inventario"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    try:
        producto = Inventario.objects.get(id=producto_id)
    except Inventario.DoesNotExist:
        messages.error(request, "El producto no existe.")
        return redirect('inventario')
    
    if request.method == 'POST':
        nombre_producto = producto.nombre
        categoria_producto = producto.categoria
        cantidad_producto = producto.cantidad
        # Registrar auditoría antes de eliminar
        from .utils import registrar_auditoria
        registrar_auditoria(
            usuario=request.user,
            accion='inventario_delete',
            modulo='inventario',
            descripcion=f'Producto "{nombre_producto}" eliminado',
            objeto_afectado=nombre_producto,
            detalles={
                'categoria': categoria_producto,
                'cantidad': cantidad_producto,
            }
        )
        # Eliminar la imagen si existe
        if producto.imagen:
            producto.imagen.delete()
        producto.delete()
        messages.success(request, f'Producto "{nombre_producto}" eliminado exitosamente.')
        return redirect('inventario')
    
    return render(request, 'core/eliminar_inventario.html', {'producto': producto})


@login_required
def productos_view(request):
    """Vista para gestionar productos (inventario con precios editables)"""
    # Todos los usuarios deben cambiar su contraseña si está requerido
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    # Obtener productos agrupados por categoría
    productos = Inventario.objects.all().order_by('categoria', 'nombre')
    bodega = productos.filter(categoria='bodega')
    meson = productos.filter(categoria='meson')
    limpieza = productos.filter(categoria='limpieza')
    
    return render(request, 'core/productos.html', {
        'bodega': bodega,
        'meson': meson,
        'limpieza': limpieza
    })


@login_required
def pedidos_view(request):
    """Vista para gestionar pedidos"""
    # Todos los usuarios deben cambiar su contraseña si está requerido
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    pedidos = Pedido.objects.all().order_by('-fecha_creacion')
    
    # Calcular el total gastado en todos los pedidos
    total_gastado = 0
    for pedido in pedidos:
        for detalle in pedido.detalles.all():
            total_gastado += detalle.cantidad * detalle.precio_unitario
    
    return render(request, 'core/pedidos.html', {
        'pedidos': pedidos,
        'total_gastado': total_gastado
    })


@login_required
def editar_precio_producto_view(request, producto_id):
    """Vista para editar solo el precio de un producto"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    producto = get_object_or_404(Inventario, id=producto_id)
    
    if request.method == 'POST':
        form = EditarPrecioProductoForm(request.POST, instance=producto)
        if form.is_valid():
            precio_anterior = producto.precio_unitario
            form.save()
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='inventario_edit',
                modulo='inventario',
                descripcion=f'Precio de producto "{producto.nombre}" modificado',
                objeto_afectado=producto.nombre,
                detalles={
                    'precio_anterior': str(precio_anterior),
                    'precio_nuevo': str(producto.precio_unitario),
                }
            )
            messages.success(request, f'Precio de "{producto.nombre}" actualizado exitosamente.')
            return redirect('productos')
    else:
        form = EditarPrecioProductoForm(instance=producto)
    
    return render(request, 'core/editar_precio_producto.html', {
        'form': form,
        'producto': producto
    })


@login_required
def agregar_al_carrito_view(request, producto_id):
    """Vista para agregar un producto al carrito"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    producto = get_object_or_404(Inventario, id=producto_id)
    
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 1))
        
        if cantidad <= 0:
            messages.error(request, 'La cantidad debe ser mayor a 0')
            return redirect('productos')
        
        # Obtener el carrito de la sesión
        carrito = request.session.get('carrito', {})
        
        # Si el producto ya está en el carrito, sumar la cantidad
        if str(producto_id) in carrito:
            carrito[str(producto_id)]['cantidad'] += cantidad
        else:
            # Agregar nuevo producto al carrito
            carrito[str(producto_id)] = {
                'nombre': producto.nombre,
                'precio': float(producto.precio_unitario),
                'cantidad': cantidad
            }
        
        # Guardar el carrito en la sesión
        request.session['carrito'] = carrito
        messages.success(request, f'"{producto.nombre}" agregado al carrito')
        return redirect('productos')
    
    return redirect('productos')


@login_required
def ver_carrito_view(request):
    """Vista para ver el carrito de compras"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    carrito = request.session.get('carrito', {})
    productos_carrito = []
    total = 0
    
    for producto_id, datos in carrito.items():
        try:
            producto = Inventario.objects.get(id=producto_id)
            subtotal = datos['precio'] * datos['cantidad']
            total += subtotal
            productos_carrito.append({
                'id': producto_id,
                'producto': producto,
                'nombre': datos['nombre'],
                'precio': datos['precio'],
                'cantidad': datos['cantidad'],
                'subtotal': subtotal
            })
        except Inventario.DoesNotExist:
            # Si el producto ya no existe, eliminarlo del carrito
            del carrito[producto_id]
            request.session['carrito'] = carrito
    
    return render(request, 'core/carrito.html', {
        'productos_carrito': productos_carrito,
        'total': total
    })


@login_required
def eliminar_del_carrito_view(request, producto_id):
    """Vista para eliminar un producto del carrito"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    carrito = request.session.get('carrito', {})
    
    if str(producto_id) in carrito:
        nombre_producto = carrito[str(producto_id)]['nombre']
        del carrito[str(producto_id)]
        request.session['carrito'] = carrito
        messages.success(request, f'"{nombre_producto}" eliminado del carrito')
    
    return redirect('ver_carrito')


@login_required
def modificar_cantidad_carrito_view(request, producto_id):
    """Vista para modificar la cantidad de un producto en el carrito"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    carrito = request.session.get('carrito', {})
    
    if request.method == 'POST':
        nueva_cantidad = int(request.POST.get('cantidad', 1))
        
        if nueva_cantidad <= 0:
            messages.error(request, 'La cantidad debe ser mayor a 0')
            return redirect('ver_carrito')
        
        if str(producto_id) in carrito:
            carrito[str(producto_id)]['cantidad'] = nueva_cantidad
            request.session['carrito'] = carrito
            messages.success(request, 'Cantidad actualizada')
    
    return redirect('ver_carrito')


@login_required
def crear_pedido_view(request):
    """Vista para crear un nuevo pedido desde el carrito"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    carrito = request.session.get('carrito', {})
    
    if not carrito:
        messages.error(request, 'El carrito está vacío. Agrega productos antes de crear un pedido.')
        return redirect('productos')
    
    # Calcular el total del carrito
    total = 0
    productos_carrito = []
    for producto_id, datos in carrito.items():
        subtotal = datos['precio'] * datos['cantidad']
        total += subtotal
        productos_carrito.append({
            'id': producto_id,
            'nombre': datos['nombre'],
            'precio': datos['precio'],
            'cantidad': datos['cantidad'],
            'subtotal': subtotal
        })
    
    if request.method == 'POST':
        form = CrearPedidoForm(request.POST)
        
        if form.is_valid():
            pedido = form.save(commit=False)
            pedido.usuario_creacion = request.user
            pedido.save()
            
            # Crear los detalles del pedido desde el carrito
            productos_data = []
            for producto_id, datos in carrito.items():
                DetallePedido.objects.create(
                    pedido=pedido,
                    producto_nombre=datos['nombre'],
                    cantidad=datos['cantidad'],
                    precio_unitario=datos['precio']
                )
                productos_data.append(datos)
            
            # Limpiar el carrito después de crear el pedido
            request.session['carrito'] = {}
            
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='pedido_create',
                modulo='pedidos',
                descripcion=f'Pedido "{pedido.codigo}" creado',
                objeto_afectado=pedido.codigo,
                detalles={
                    'codigo': pedido.codigo,
                    'productos': len(productos_data),
                }
            )
            
            messages.success(request, f'Pedido "{pedido.codigo}" creado exitosamente.')
            return redirect('pedidos')
    else:
        form = CrearPedidoForm()
    
    return render(request, 'core/crear_pedido.html', {
        'form': form,
        'carrito': carrito,
        'productos_carrito': productos_carrito,
        'total': total
    })


@login_required
def exportar_pedido_excel(request, pedido_id):
    """Vista para exportar un pedido a Excel"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        messages.error(request, 'La librería openpyxl no está instalada.')
        return redirect('pedidos')
    
    pedido = get_object_or_404(Pedido, id=pedido_id)
    detalles = pedido.detalles.all()
    
    # Crear un libro de trabajo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido"
    
    # Estilos
    header_fill = PatternFill(start_color="B08968", end_color="B08968", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    
    # Encabezado
    ws['A1'] = "PEDIDO"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Código: {pedido.codigo}"
    ws['A3'] = f"Fecha: {pedido.fecha_creacion.strftime('%d/%m/%Y %H:%M')}"
    ws['A4'] = f"Usuario: {pedido.usuario_creacion.get_nombre_completo() if pedido.usuario_creacion else 'N/A'}"
    
    if pedido.observaciones:
        ws['A5'] = f"Observaciones: {pedido.observaciones}"
    
    # Encabezados de tabla
    ws['A7'] = "Nombre del Producto"
    ws['B7'] = "Cantidad"
    ws['C7'] = "Precio Unitario"
    ws['D7'] = "Total"
    
    for cell in ['A7', 'B7', 'C7', 'D7']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    row = 8
    total_general = 0
    for detalle in detalles:
        ws[f'A{row}'] = detalle.producto_nombre
        ws[f'B{row}'] = detalle.cantidad
        ws[f'C{row}'] = float(detalle.precio_unitario)
        ws[f'C{row}'].number_format = '#,##0.00'
        total_detalle = detalle.precio_total
        total_general += total_detalle
        ws[f'D{row}'] = float(total_detalle)
        ws[f'D{row}'].number_format = '#,##0.00'
        row += 1
    
    # Fila de total
    ws[f'A{row}'] = "TOTAL"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'D{row}'] = float(total_general)
    ws[f'D{row}'].font = Font(bold=True, size=12)
    ws[f'D{row}'].number_format = '#,##0.00'
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    
    # Respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = f"Pedido_{pedido.codigo.replace(' ', '_')}_{pedido.fecha_creacion.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    
    wb.save(response)
    return response


@login_required
def eliminar_pedido_view(request, pedido_id):
    """Vista para eliminar un pedido"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    if request.method == 'POST':
        codigo_pedido = pedido.codigo
        pedido.delete()
        messages.success(request, f'Pedido "{codigo_pedido}" eliminado exitosamente.')
        return redirect('pedidos')
    
    return render(request, 'core/eliminar_pedido.html', {'pedido': pedido})


@login_required
def deliverys_view(request):
    """Vista para gestionar deliverys"""
    # Todos los usuarios deben cambiar su contraseña si está requerido
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    # Por ahora, renderizar una página básica
    return render(request, 'core/deliverys.html')


@login_required
def operaciones_view(request):
    """Vista para gestionar operaciones"""
    # Todos los usuarios deben cambiar su contraseña si está requerido
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    from datetime import datetime, timedelta
    
    # Obtener registros de fallas y llamadas
    fallas = RegistroFalla.objects.all().order_by('-fecha', '-contador_falla')
    llamadas = RegistroLlamada.objects.all().order_by('-fecha', '-contador_llamada')
    
    # Calcular porcentajes mensuales
    fecha_actual = timezone.localtime(timezone.now())
    inicio_mes = fecha_actual.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Calcular días transcurridos del mes actual
    dias_transcurridos = fecha_actual.day
    
    # Fallas del mes actual
    fallas_mes = RegistroFalla.objects.filter(fecha__gte=inicio_mes)
    total_fallas_mes = fallas_mes.count()
    
    # Calcular días únicos con fallas
    dias_con_fallas = fallas_mes.values('fecha').distinct().count()
    porcentaje_fallas_mes = round((dias_con_fallas / dias_transcurridos * 100), 1) if dias_transcurridos > 0 else 0
    
    # Llamadas del mes actual
    llamadas_mes = RegistroLlamada.objects.filter(fecha__gte=inicio_mes)
    total_llamadas_mes = llamadas_mes.count()
    
    # Calcular días únicos con llamadas
    dias_con_llamadas = llamadas_mes.values('fecha').distinct().count()
    porcentaje_llamadas_mes = round((dias_con_llamadas / dias_transcurridos * 100), 1) if dias_transcurridos > 0 else 0
    
    # Obtener siguiente contador para fallas y llamadas
    siguiente_contador_falla = (RegistroFalla.objects.aggregate(Max('contador_falla'))['contador_falla__max'] or 0) + 1
    siguiente_contador_llamada = (RegistroLlamada.objects.aggregate(Max('contador_llamada'))['contador_llamada__max'] or 0) + 1
    
    return render(request, 'core/operaciones.html', {
        'fallas': fallas[:10],  # Últimas 10 fallas
        'llamadas': llamadas[:10],  # Últimas 10 llamadas
        'total_fallas_mes': total_fallas_mes,
        'total_llamadas_mes': total_llamadas_mes,
        'porcentaje_fallas_mes': porcentaje_fallas_mes,
        'porcentaje_llamadas_mes': porcentaje_llamadas_mes,
        'siguiente_contador_falla': siguiente_contador_falla,
        'siguiente_contador_llamada': siguiente_contador_llamada,
    })


@login_required
def crear_falla_view(request):
    """Vista para crear un nuevo registro de falla"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    
    if request.method == 'POST':
        form = RegistroFallaForm(request.POST)
        if form.is_valid():
            falla = form.save(commit=False)
            # Asignar contador automáticamente
            ultimo_contador = RegistroFalla.objects.aggregate(Max('contador_falla'))['contador_falla__max']
            falla.contador_falla = (ultimo_contador or 0) + 1
            falla.usuario_registro = request.user
            falla.save()
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='falla_create',
                modulo='operaciones',
                descripcion=f'Falla #{falla.contador_falla} registrada: {falla.maquina}',
                objeto_afectado=f"Falla #{falla.contador_falla}",
                detalles={
                    'contador_falla': falla.contador_falla,
                    'maquina': falla.maquina,
                    'fecha': str(falla.fecha),
                }
            )
            messages.success(request, f'Falla #{falla.contador_falla} registrada exitosamente.')
            return redirect('operaciones')
    else:
        form = RegistroFallaForm()
    
    return render(request, 'core/crear_falla.html', {'form': form})


@login_required
def crear_llamada_view(request):
    """Vista para crear un nuevo registro de llamada"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    
    if request.method == 'POST':
        form = RegistroLlamadaForm(request.POST)
        if form.is_valid():
            llamada = form.save(commit=False)
            # Asignar contador automáticamente
            ultimo_contador = RegistroLlamada.objects.aggregate(Max('contador_llamada'))['contador_llamada__max']
            llamada.contador_llamada = (ultimo_contador or 0) + 1
            llamada.usuario_registro = request.user
            llamada.save()
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='llamada_create',
                modulo='operaciones',
                descripcion=f'Llamada #{llamada.contador_llamada} registrada: {llamada.motivo}',
                objeto_afectado=f"Llamada #{llamada.contador_llamada}",
                detalles={
                    'contador_llamada': llamada.contador_llamada,
                    'motivo': llamada.motivo,
                    'tecnico_contactado': llamada.tecnico_contactado,
                    'fecha': str(llamada.fecha),
                }
            )
            messages.success(request, f'Llamada #{llamada.contador_llamada} registrada exitosamente.')
            return redirect('operaciones')
    else:
        form = RegistroLlamadaForm()
    
    return render(request, 'core/crear_llamada.html', {'form': form})


@login_required
def editar_falla_view(request, falla_id):
    """Vista para editar un registro de falla"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    falla = get_object_or_404(RegistroFalla, id=falla_id)
    
    if request.method == 'POST':
        form = RegistroFallaForm(request.POST, instance=falla)
        if form.is_valid():
            form.save()
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='falla_edit',
                modulo='operaciones',
                descripcion=f'Falla #{falla.contador_falla} modificada: {falla.maquina}',
                objeto_afectado=f"Falla #{falla.contador_falla}",
                detalles={
                    'contador_falla': falla.contador_falla,
                    'maquina': falla.maquina,
                    'fecha': str(falla.fecha),
                }
            )
            messages.success(request, f'Falla #{falla.contador_falla} actualizada exitosamente.')
            return redirect('operaciones')
    else:
        form = RegistroFallaForm(instance=falla)
    
    return render(request, 'core/editar_falla.html', {'form': form, 'falla': falla})


@login_required
def eliminar_falla_view(request, falla_id):
    """Vista para eliminar un registro de falla"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    falla = get_object_or_404(RegistroFalla, id=falla_id)
    
    if request.method == 'POST':
        contador = falla.contador_falla
        maquina = falla.maquina
        fecha = str(falla.fecha)
        falla.delete()
        # Registrar auditoría
        from .utils import registrar_auditoria
        registrar_auditoria(
            usuario=request.user,
            accion='falla_delete',
            modulo='operaciones',
                descripcion=f'Falla #{contador} eliminada: {maquina}',
            objeto_afectado=f"Falla #{contador}",
            detalles={
                'contador_falla': contador,
                'maquina': maquina,
                'fecha': fecha,
            }
        )
        messages.success(request, f'Falla #{contador} eliminada exitosamente.')
        return redirect('operaciones')
    
    return render(request, 'core/eliminar_falla.html', {'falla': falla})


@login_required
def editar_llamada_view(request, llamada_id):
    """Vista para editar un registro de llamada"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    llamada = get_object_or_404(RegistroLlamada, id=llamada_id)
    
    if request.method == 'POST':
        form = RegistroLlamadaForm(request.POST, instance=llamada)
        if form.is_valid():
            form.save()
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='llamada_edit',
                modulo='operaciones',
                descripcion=f'Llamada #{llamada.contador_llamada} modificada: {llamada.motivo}',
                objeto_afectado=f"Llamada #{llamada.contador_llamada}",
                detalles={
                    'contador_llamada': llamada.contador_llamada,
                    'motivo': llamada.motivo,
                    'tecnico_contactado': llamada.tecnico_contactado,
                    'fecha': str(llamada.fecha),
                }
            )
            messages.success(request, f'Llamada #{llamada.contador_llamada} actualizada exitosamente.')
            return redirect('operaciones')
    else:
        form = RegistroLlamadaForm(instance=llamada)
    
    return render(request, 'core/editar_llamada.html', {'form': form, 'llamada': llamada})


@login_required
def eliminar_llamada_view(request, llamada_id):
    """Vista para eliminar un registro de llamada"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    llamada = get_object_or_404(RegistroLlamada, id=llamada_id)
    
    if request.method == 'POST':
        contador = llamada.contador_llamada
        motivo = llamada.motivo
        tecnico = llamada.tecnico_contactado
        fecha = str(llamada.fecha)
        llamada.delete()
        # Registrar auditoría
        from .utils import registrar_auditoria
        registrar_auditoria(
            usuario=request.user,
            accion='llamada_delete',
            modulo='operaciones',
                descripcion=f'Llamada #{contador} eliminada: {motivo}',
            objeto_afectado=f"Llamada #{contador}",
            detalles={
                'contador_llamada': contador,
                'motivo': motivo,
                'tecnico_contactado': tecnico,
                'fecha': fecha,
            }
        )
        messages.success(request, f'Llamada #{contador} eliminada exitosamente.')
        return redirect('operaciones')
    
    return render(request, 'core/eliminar_llamada.html', {'llamada': llamada})


@login_required
def asistencia_view(request):
    """Vista para gestionar asistencia"""
    # Todos los usuarios deben cambiar su contraseña si está requerido
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    from datetime import datetime, timedelta
    
    # Obtener parámetros de filtrado
    busqueda = request.GET.get('busqueda', '').strip()
    orden = request.GET.get('orden', 'desc')  # 'asc' o 'desc'
    filtro_fecha = request.GET.get('filtro_fecha', '')  # 'hoy', 'semana', 'mes', 'todos'
    filtro_turno = request.GET.get('filtro_turno', '')  # 'apertura', 'tarde', 'cierre', ''
    
    # Iniciar con todas las asistencias
    asistencias = Asistencia.objects.select_related('usuario').all()
    
    # Aplicar búsqueda por nombre (coincidencias parciales)
    if busqueda:
        # Buscar en nombre, apellido, nombre completo y username
        # El operador icontains busca coincidencias parciales sin distinguir mayúsculas/minúsculas
        asistencias = asistencias.filter(
            Q(usuario__nombre__icontains=busqueda) |
            Q(usuario__apellido__icontains=busqueda) |
            Q(usuario__username__icontains=busqueda)
        )
    
    # Aplicar filtro por fecha
    fecha_actual = timezone.localtime(timezone.now()).date()
    if filtro_fecha == 'hoy':
        asistencias = asistencias.filter(fecha=fecha_actual)
    elif filtro_fecha == 'semana':
        fecha_inicio = fecha_actual - timedelta(days=7)
        asistencias = asistencias.filter(fecha__gte=fecha_inicio)
    elif filtro_fecha == 'mes':
        fecha_inicio = fecha_actual - timedelta(days=30)
        asistencias = asistencias.filter(fecha__gte=fecha_inicio)
    # Si es 'todos' o vacío, no se filtra por fecha
    
    # Aplicar filtro por turno
    if filtro_turno:
        asistencias = asistencias.filter(turno=filtro_turno)
    
    # Aplicar ordenamiento
    if orden == 'asc':
        asistencias = asistencias.order_by('fecha', 'fecha_registro')
    else:  # desc (por defecto)
        asistencias = asistencias.order_by('-fecha', '-fecha_registro')
    
    # Calcular usuario con mayor y menor asistencia
    usuarios_asistencias = Usuario.objects.annotate(
        total_asistencias=Count('asistencias')
    ).filter(total_asistencias__gt=0).order_by('-total_asistencias')
    
    usuario_mayor_asistencia = usuarios_asistencias.first() if usuarios_asistencias.exists() else None
    usuario_menor_asistencia = usuarios_asistencias.last() if usuarios_asistencias.exists() else None
    
    # Si hay solo un usuario, no hay menor (es el mismo)
    if usuarios_asistencias.count() == 1:
        usuario_menor_asistencia = None
    
    return render(request, 'core/asistencia.html', {
        'asistencias': asistencias,
        'busqueda': busqueda,
        'orden': orden,
        'filtro_fecha': filtro_fecha,
        'filtro_turno': filtro_turno,
        'usuario_mayor_asistencia': usuario_mayor_asistencia,
        'usuario_menor_asistencia': usuario_menor_asistencia,
    })


@login_required
def editar_asistencia_view(request, asistencia_id):
    """Vista para editar un registro de asistencia"""
    # Todos los usuarios deben cambiar su contraseña si está requerido
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    asistencia = get_object_or_404(Asistencia, id=asistencia_id)
    
    if request.method == 'POST':
        form = EditarAsistenciaForm(request.POST, instance=asistencia)
        if form.is_valid():
            usuario_nombre = asistencia.usuario.get_nombre_completo() if asistencia.usuario else 'Usuario Anónimo'
            usuario_nombre_completo = f"{asistencia.usuario.nombre} {asistencia.usuario.apellido}" if asistencia.usuario else 'Usuario Anónimo'
            form.save()
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='asistencia_edit',
                modulo='asistencia',
                descripcion=f'Asistencia modificada de {usuario_nombre} - Fecha: {asistencia.fecha}, Turno: {asistencia.get_turno_display()}',
                objeto_afectado=usuario_nombre_completo,
                detalles={
                    'usuario': usuario_nombre,
                    'fecha': str(asistencia.fecha),
                    'turno': asistencia.turno,
                }
            )
            messages.success(request, 'Registro de asistencia actualizado exitosamente.')
            return redirect('asistencia')
    else:
        form = EditarAsistenciaForm(instance=asistencia)
    
    return render(request, 'core/editar_asistencia.html', {'form': form, 'asistencia': asistencia})


@login_required
def eliminar_asistencia_view(request, asistencia_id):
    """Vista para eliminar un registro de asistencia"""
    # Todos los usuarios deben cambiar su contraseña si está requerido
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    asistencia = get_object_or_404(Asistencia, id=asistencia_id)
    
    if request.method == 'POST':
        usuario_nombre = asistencia.usuario.get_nombre_completo() if asistencia.usuario else 'Usuario Anónimo'
        usuario_nombre_completo = f"{asistencia.usuario.nombre} {asistencia.usuario.apellido}" if asistencia.usuario else 'Usuario Anónimo'
        fecha = str(asistencia.fecha)
        turno = asistencia.turno
        turno_display = asistencia.get_turno_display()
        asistencia.delete()
        # Registrar auditoría
        from .utils import registrar_auditoria
        registrar_auditoria(
            usuario=request.user,
            accion='asistencia_delete',
            modulo='asistencia',
                descripcion=f'Asistencia eliminada de {usuario_nombre} - Fecha: {fecha}, Turno: {turno_display}',
            objeto_afectado=usuario_nombre_completo,
            detalles={
                'usuario': usuario_nombre,
                'fecha': fecha,
                'turno': turno,
            }
        )
        messages.success(request, f'Registro de asistencia de {usuario_nombre} eliminado exitosamente.')
        return redirect('asistencia')
    
    return render(request, 'core/eliminar_asistencia.html', {'asistencia': asistencia})


def index_view(request):
    """Vista para la página principal"""
    # Obtener imágenes activas del carrusel ordenadas por orden
    imagenes_carrusel = list(ImagenCarrusel.objects.filter(activo=True).order_by('orden', 'fecha_creacion'))
    
    # Crear una lista de imágenes - solo mostrar las imágenes reales disponibles
    imagenes_completas = []
    
    # Agregar las imágenes reales ordenadas por su campo 'orden'
    for imagen in imagenes_carrusel:
        imagenes_completas.append({
            'tipo': 'real',
            'imagen': imagen,
            'indice': imagen.orden if imagen.orden > 0 else len(imagenes_completas) + 1,
            'titulo_barista': imagen.titulo_barista if imagen.orden == 1 else None
        })
    
    # Ordenar por índice (orden) para asegurar el orden correcto
    imagenes_completas.sort(key=lambda x: x['indice'])
    
    # Obtener eventos activos (ordenados por fecha del evento, más recientes primero)
    eventos = Evento.objects.filter(activo=True).order_by('-fecha_evento')[:1]  # Solo el más reciente
    
    # Obtener noticias activas (ordenadas por fecha de publicación, más recientes primero)
    noticias = Noticia.objects.filter(activo=True).order_by('-fecha_publicacion')[:1]  # Solo la más reciente
    
    return render(request, 'core/index.html', {
        'imagenes_carrusel': imagenes_completas,
        'eventos': eventos,
        'noticias': noticias
    })


def documentacion_view(request):
    """Vista para mostrar la documentación (reglamento interno y manual de operaciones)"""
    # Obtener los documentos activos más recientes de cada tipo
    reglamento = ManualInterno.objects.filter(activo=True, tipo='reglamento').order_by('-fecha_actualizacion').first()
    manual_operaciones = ManualInterno.objects.filter(activo=True, tipo='operaciones').order_by('-fecha_actualizacion').first()
    
    return render(request, 'core/documentacion.html', {
        'reglamento': reglamento,
        'manual_operaciones': manual_operaciones
    })


def reglamento_interno_view(request):
    """Vista para mostrar o descargar el reglamento interno (mantener compatibilidad)"""
    # Obtener el reglamento interno activo más reciente
    try:
        manual = ManualInterno.objects.filter(activo=True, tipo='reglamento').order_by('-fecha_actualizacion').first()
        
        if not manual:
            messages.info(request, 'No hay un reglamento interno disponible actualmente.')
            return redirect('index')
        
        # Redirigir a la URL del archivo para descargarlo/visualizarlo
        return redirect(manual.archivo.url)
    except Exception as e:
        messages.info(request, 'No hay un reglamento interno disponible actualmente.')
        return redirect('index')


def inicializar_contactos_default():
    """Función para inicializar los contactos por defecto si no existen"""
    contactos_default = [
        {
            'nombre': 'Gerente General',
            'cargo': 'Gerente General',
            'email': 'gerente.general@popupnescafe.cl',
            'telefono': '+56 9 XXXX XXXX',
            'orden': 1,
        },
        {
            'nombre': 'Gerente de Operaciones',
            'cargo': 'Gerente de Operaciones',
            'email': 'operaciones@popupnescafe.cl',
            'telefono': '+56 9 XXXX XXXX',
            'orden': 2,
        },
        {
            'nombre': 'Gerente de Recursos Humanos',
            'cargo': 'Gerente de RRHH',
            'email': 'rrhh@popupnescafe.cl',
            'telefono': '+56 9 XXXX XXXX',
            'orden': 3,
        },
        {
            'nombre': 'Gerente de Marketing',
            'cargo': 'Gerente de Marketing',
            'email': 'marketing@popupnescafe.cl',
            'telefono': '+56 9 XXXX XXXX',
            'orden': 4,
        },
        {
            'nombre': 'Gerente de Finanzas',
            'cargo': 'Gerente de Finanzas',
            'email': 'finanzas@popupnescafe.cl',
            'telefono': '+56 9 XXXX XXXX',
            'orden': 5,
        }
    ]
    
    # Verificar si ya existen contactos
    if Contacto.objects.count() == 0:
        for contacto_data in contactos_default:
            Contacto.objects.create(
                nombre=contacto_data['nombre'],
                cargo=contacto_data['cargo'],
                email=contacto_data['email'],
                telefono=contacto_data['telefono'],
                orden=contacto_data['orden'],
                activo=True
            )


def contactanos_view(request):
    """Vista para la página de contáctanos"""
    # Verificar si el usuario está autenticado
    usuario_autenticado = request.user.is_authenticated
    
    # Inicializar contactos por defecto si no existen
    inicializar_contactos_default()
    
    # Obtener contactos activos de la base de datos
    try:
        contactos = Contacto.objects.filter(activo=True).order_by('orden', 'nombre')
    except:
        # Si hay error, usar lista vacía (el template manejará el caso vacío)
        contactos = []
    
    return render(request, 'core/contactanos.html', {
        'contactos': contactos,
        'usuario_autenticado': usuario_autenticado
    })


def equipo_view(request):
    """Vista para la página de conoce al equipo"""
    return render(request, 'core/equipo.html')


def creditos_view(request):
    """Vista para la página de créditos y agradecimientos"""
    return render(request, 'core/creditos.html')


@login_required
def gestionar_contactos_view(request):
    """Vista para gestionar contactos"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    # Inicializar contactos por defecto si no existen
    inicializar_contactos_default()
    
    contactos = Contacto.objects.all().order_by('orden', 'nombre')
    return render(request, 'core/gestionar_contactos.html', {'contactos': contactos})


@login_required
def editar_contacto_view(request, contacto_id):
    """Vista para editar un contacto"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    contacto = get_object_or_404(Contacto, id=contacto_id)
    
    if request.method == 'POST':
        form = ContactoForm(request.POST, request.FILES, instance=contacto)
        if form.is_valid():
            # Manejar eliminación de imagen
            if form.cleaned_data.get('eliminar_imagen'):
                if contacto.imagen:
                    if default_storage.exists(contacto.imagen.name):
                        default_storage.delete(contacto.imagen.name)
                    contacto.imagen = None
            
            form.save()
            contacto.refresh_from_db()
            
            # Si se marcó eliminar pero no se subió nueva imagen, asegurar que se elimine
            if form.cleaned_data.get('eliminar_imagen') and not form.cleaned_data.get('imagen'):
                Contacto.objects.filter(id=contacto.id).update(imagen=None)
                contacto.refresh_from_db()
            
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='contacto_edit',
                modulo='contenido',
                descripcion=f'Contacto "{contacto.nombre}" modificado',
                objeto_afectado=contacto.nombre,
                detalles={
                    'nombre': contacto.nombre,
                    'cargo': contacto.cargo,
                }
            )
            
            if form.cleaned_data.get('eliminar_imagen'):
                messages.success(request, f'Contacto "{contacto.nombre}" actualizado exitosamente. Imagen eliminada.')
            else:
                messages.success(request, f'Contacto "{contacto.nombre}" actualizado exitosamente.')
            return redirect('gestionar_contactos')
    else:
        form = ContactoForm(instance=contacto)
    
    return render(request, 'core/editar_contacto.html', {
        'form': form,
        'contacto': contacto
    })


@login_required
def crear_contacto_view(request):
    """Vista para crear un nuevo contacto"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    if request.method == 'POST':
        form = ContactoForm(request.POST, request.FILES)
        if form.is_valid():
            contacto = form.save()
            
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='contacto_create',
                modulo='contenido',
                descripcion=f'Contacto "{contacto.nombre}" creado',
                objeto_afectado=contacto.nombre,
                detalles={
                    'nombre': contacto.nombre,
                    'cargo': contacto.cargo,
                }
            )
            
            messages.success(request, f'Contacto "{contacto.nombre}" creado exitosamente.')
            return redirect('gestionar_contactos')
    else:
        form = ContactoForm()
    
    return render(request, 'core/crear_contacto.html', {'form': form})


@login_required
def eliminar_contacto_view(request, contacto_id):
    """Vista para eliminar un contacto"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    contacto = get_object_or_404(Contacto, id=contacto_id)
    
    if request.method == 'POST':
        nombre_contacto = contacto.nombre
        # Eliminar imagen si existe
        if contacto.imagen:
            if default_storage.exists(contacto.imagen.name):
                default_storage.delete(contacto.imagen.name)
        contacto.delete()
        
        # Registrar auditoría
        from .utils import registrar_auditoria
        registrar_auditoria(
            usuario=request.user,
            accion='contacto_delete',
            modulo='contenido',
            descripcion=f'Contacto "{nombre_contacto}" eliminado',
            objeto_afectado=nombre_contacto,
            detalles={}
        )
        
        messages.success(request, f'Contacto "{nombre_contacto}" eliminado exitosamente.')
        return redirect('gestionar_contactos')
    
    return render(request, 'core/eliminar_contacto.html', {'contacto': contacto})


@login_required
def gestionar_manual_interno_view(request):
    """Vista para gestionar documentación"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    # Filtrar por tipo si se proporciona
    tipo_filtro = request.GET.get('tipo', '')
    manuales = ManualInterno.objects.all()
    
    if tipo_filtro:
        manuales = manuales.filter(tipo=tipo_filtro)
    
    manuales = manuales.order_by('tipo', '-fecha_actualizacion')
    return render(request, 'core/gestionar_manual_interno.html', {
        'manuales': manuales,
        'tipo_filtro': tipo_filtro
    })


@login_required
def crear_manual_interno_view(request):
    """Vista para crear un nuevo manual interno"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    if request.method == 'POST':
        form = ManualInternoForm(request.POST, request.FILES)
        if form.is_valid():
            manual = form.save()
            
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='manual_create',
                modulo='contenido',
                descripcion=f'Documento "{manual.get_tipo_display()}" - "{manual.titulo}" creado',
                objeto_afectado=manual.titulo,
                detalles={
                    'titulo': manual.titulo,
                    'tipo': manual.get_tipo_display(),
                }
            )
            
            messages.success(request, f'Documento "{manual.get_tipo_display()}" - "{manual.titulo}" creado exitosamente.')
            return redirect('gestionar_manual_interno')
    else:
        form = ManualInternoForm()
    
    return render(request, 'core/crear_manual_interno.html', {'form': form})


@login_required
def editar_manual_interno_view(request, manual_id):
    """Vista para editar un manual interno"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    manual = get_object_or_404(ManualInterno, id=manual_id)
    
    if request.method == 'POST':
        # Guardar archivo anterior si existe
        archivo_anterior = manual.archivo.name if manual.archivo else None
        
        form = ManualInternoForm(request.POST, request.FILES, instance=manual)
        if form.is_valid():
            manual = form.save()
            
            # Si se subió un nuevo archivo y había uno anterior, eliminar el anterior
            if request.FILES.get('archivo') and archivo_anterior:
                if default_storage.exists(archivo_anterior):
                    default_storage.delete(archivo_anterior)
            
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='manual_edit',
                modulo='contenido',
                descripcion=f'Documento "{manual.get_tipo_display()}" - "{manual.titulo}" modificado',
                objeto_afectado=manual.titulo,
                detalles={
                    'titulo': manual.titulo,
                    'tipo': manual.get_tipo_display(),
                }
            )
            
            messages.success(request, f'Documento "{manual.get_tipo_display()}" - "{manual.titulo}" actualizado exitosamente.')
            return redirect('gestionar_manual_interno')
    else:
        form = ManualInternoForm(instance=manual)
    
    return render(request, 'core/editar_manual_interno.html', {
        'form': form,
        'manual': manual
    })


@login_required
def eliminar_manual_interno_view(request, manual_id):
    """Vista para eliminar un manual interno"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    manual = get_object_or_404(ManualInterno, id=manual_id)
    
    if request.method == 'POST':
        titulo_manual = manual.titulo
        # Eliminar archivo si existe
        if manual.archivo:
            if default_storage.exists(manual.archivo.name):
                default_storage.delete(manual.archivo.name)
        manual.delete()
        
        # Registrar auditoría
        from .utils import registrar_auditoria
        tipo_manual = manual.get_tipo_display()
        registrar_auditoria(
            usuario=request.user,
            accion='manual_delete',
            modulo='contenido',
            descripcion=f'Documento "{tipo_manual}" - "{titulo_manual}" eliminado',
            objeto_afectado=titulo_manual,
            detalles={
                'tipo': tipo_manual,
            }
        )
        
        messages.success(request, f'Documento "{tipo_manual}" - "{titulo_manual}" eliminado exitosamente.')
        return redirect('gestionar_manual_interno')
    
    return render(request, 'core/eliminar_manual_interno.html', {'manual': manual})


@login_required
def colaborador_dashboard_view(request):
    """Dashboard principal para usuarios colaboradores"""
    # Solo usuarios colaboradores pueden acceder (pero si es admin también puede acceder)
    if not request.user.es_colaborador and not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    # Si el usuario necesita cambiar su contraseña, redirigir primero
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    
    return render(request, 'core/colaborador_dashboard.html', {
        'usuario': request.user
    })


@login_required
def registro_asistencia_view(request):
    """Vista para registro de asistencia (solo usuarios colaboradores)"""
    # Solo usuarios colaboradores pueden registrar asistencia
    if not request.user.es_colaborador:
        messages.error(request, 'Solo los usuarios colaboradores pueden registrar asistencia')
        return redirect('panel')
    
    # Si el usuario necesita cambiar su contraseña, redirigir primero
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    
    form = RegistroAsistenciaForm()
    
    if request.method == 'POST':
        form = RegistroAsistenciaForm(request.POST)
        if form.is_valid():
            turno = form.cleaned_data['turno']
            
            # Obtener el usuario logueado
            usuario = request.user
            
            # Mapear turno a horas de entrada y salida (horario chileno GMT-3)
            horarios_turno = {
                'apertura': {'entrada': '09:00', 'salida': '13:00'},
                'tarde': {'entrada': '13:00', 'salida': '17:00'},
                'cierre': {'entrada': '17:00', 'salida': '21:00'}
            }
            horario = horarios_turno.get(turno, {'entrada': '09:00', 'salida': '13:00'})
            from datetime import datetime
            hora_entrada = datetime.strptime(horario['entrada'], '%H:%M').time()
            hora_salida = datetime.strptime(horario['salida'], '%H:%M').time()
            
            try:
                # Verificar si ya existe asistencia para hoy (usando zona horaria de Chile GMT-3)
                fecha_hoy = timezone.localtime(timezone.now()).date()
                asistencia_existente = Asistencia.objects.filter(
                    usuario=usuario,
                    fecha=fecha_hoy
                ).first()
                
                if asistencia_existente:
                    messages.warning(request, f'Ya se registró asistencia para hoy. Tu turno registrado es: {asistencia_existente.get_turno_display()}')
                    # Redirigir al dashboard del colaborador
                    return redirect('colaborador_dashboard')
                else:
                    # Crear nuevo registro de asistencia
                    # fecha_registro se guarda automáticamente con la fecha y hora exacta del registro (GMT-3 Chile)
                    fecha_registro_chile = timezone.localtime(timezone.now())
                    asistencia = Asistencia.objects.create(
                        usuario=usuario,
                        fecha=fecha_hoy,
                        hora_entrada=hora_entrada,
                        hora_salida=hora_salida,
                        estado='presente',
                        turno=turno,
                        fecha_registro=fecha_registro_chile  # Fecha y hora exacta del registro en horario chileno GMT-3
                    )
                    messages.success(request, f'Asistencia registrada exitosamente. Turno: {asistencia.get_turno_display()}')
                    # Redirigir al dashboard del colaborador
                    return redirect('colaborador_dashboard')
                
            except Exception as e:
                messages.error(request, f'Error al registrar asistencia: {str(e)}')
                logger.error(f'Error al registrar asistencia: {traceback.format_exc()}')
    
    # Obtener la fecha actual para mostrar en el template (horario chileno GMT-3)
    fecha_actual = timezone.localtime(timezone.now()).date()
    return render(request, 'core/registro_asistencia.html', {
        'form': form, 
        'usuario': request.user,
        'fecha_actual': fecha_actual
    })


@login_required
def cambiar_stock_view(request):
    """Vista para que colaboradores cambien el stock de productos"""
    # Solo usuarios colaboradores pueden cambiar stock
    if not request.user.es_colaborador:
        messages.error(request, 'Solo los usuarios colaboradores pueden cambiar el stock')
        return redirect('panel')
    
    # Si el usuario necesita cambiar su contraseña, redirigir primero
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    
    # Obtener todos los productos
    productos = Inventario.objects.all().order_by('categoria', 'nombre')
    
    # Organizar productos por categoría
    bodega = productos.filter(categoria='bodega')
    meson = productos.filter(categoria='meson')
    limpieza = productos.filter(categoria='limpieza')
    
    return render(request, 'core/cambiar_stock.html', {
        'bodega': bodega,
        'meson': meson,
        'limpieza': limpieza,
        'usuario': request.user
    })


@login_required
def actualizar_stock_view(request, producto_id):
    """Vista para actualizar el stock de un producto específico (solo colaboradores)"""
    # Solo usuarios colaboradores pueden cambiar stock
    if not request.user.es_colaborador:
        messages.error(request, 'Solo los usuarios colaboradores pueden cambiar el stock')
        return redirect('panel')
    
    # Si el usuario necesita cambiar su contraseña, redirigir primero
    if request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    
    try:
        producto = Inventario.objects.get(id=producto_id)
    except Inventario.DoesNotExist:
        messages.error(request, 'Producto no encontrado')
        return redirect('cambiar_stock')
    
    form = CambiarStockForm(initial={'cantidad': producto.cantidad})
    
    if request.method == 'POST':
        form = CambiarStockForm(request.POST)
        if form.is_valid():
            cantidad_anterior = producto.cantidad
            nueva_cantidad = form.cleaned_data['cantidad']
            producto.cantidad = nueva_cantidad
            producto.save()
            # Registrar auditoría de cambio de stock
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='inventario_stock_change',
                modulo='inventario',
                descripcion=f'Stock de "{producto.nombre}" cambió de {cantidad_anterior} a {nueva_cantidad} unidades',
                objeto_afectado=producto.nombre,
                detalles={
                    'cantidad_anterior': cantidad_anterior,
                    'cantidad_nueva': nueva_cantidad,
                    'diferencia': nueva_cantidad - cantidad_anterior,
                    'categoria': producto.categoria,
                }
            )
            messages.success(request, f'Stock de "{producto.nombre}" actualizado a {nueva_cantidad} unidades')
            return redirect('cambiar_stock')
    
    return render(request, 'core/actualizar_stock.html', {
        'producto': producto,
        'form': form,
        'usuario': request.user
    })


@login_required
def auditoria_view(request):
    """Vista para ver los registros de auditoría"""
    # Solo administradores pueden ver la auditoría
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    # Obtener todos los registros de auditoría
    registros = Auditoria.objects.all().select_related('usuario')
    
    # Filtros
    modulo_filter = request.GET.get('modulo', '')
    accion_filter = request.GET.get('accion', '')
    usuario_filter = request.GET.get('usuario', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    search_query = request.GET.get('search', '')
    
    # Aplicar filtros
    if modulo_filter:
        registros = registros.filter(modulo=modulo_filter)
    if accion_filter:
        registros = registros.filter(accion=accion_filter)
    if usuario_filter:
        registros = registros.filter(usuario_id=usuario_filter)
    if fecha_desde:
        registros = registros.filter(fecha_hora__date__gte=fecha_desde)
    if fecha_hasta:
        registros = registros.filter(fecha_hora__date__lte=fecha_hasta)
    if search_query:
        registros = registros.filter(
            Q(descripcion__icontains=search_query) |
            Q(objeto_afectado__icontains=search_query) |
            Q(usuario__username__icontains=search_query) |
            Q(usuario__nombre__icontains=search_query) |
            Q(usuario__apellido__icontains=search_query)
        )
    
    # Ordenar por fecha más reciente
    registros = registros.order_by('-fecha_hora')
    
    # Paginación (opcional, por ahora mostramos todos)
    # registros = registros[:100]  # Limitar a los últimos 100 registros
    
    # Obtener opciones para los filtros
    usuarios = Usuario.objects.all().order_by('username')
    
    context = {
        'registros': registros,
        'usuarios': usuarios,
        'modulos': Auditoria.MODULO_CHOICES,
        'acciones': Auditoria.ACCION_CHOICES,
        'modulo_filter': modulo_filter,
        'accion_filter': accion_filter,
        'usuario_filter': usuario_filter,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'search_query': search_query,
    }
    
    return render(request, 'core/auditoria.html', context)


# ==================== GESTIÓN DE CARRUSEL ====================

@login_required
def gestionar_carrusel_view(request):
    """Vista para gestionar las imágenes del carrusel"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    imagenes = ImagenCarrusel.objects.all().order_by('orden', 'fecha_creacion')
    return render(request, 'core/gestionar_carrusel.html', {'imagenes': imagenes})


@login_required
def crear_imagen_carrusel_view(request):
    """Vista para crear una nueva imagen del carrusel"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    if request.method == 'POST':
        form = ImagenCarruselForm(request.POST, request.FILES)
        if form.is_valid():
            imagen = form.save()
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='carrusel_create',
                modulo='contenido',
                descripcion=f'Imagen de carrusel creada (orden: {imagen.orden})',
                objeto_afectado=f"Imagen Carrusel #{imagen.orden}",
                detalles={
                    'orden': imagen.orden,
                    'activo': imagen.activo,
                }
            )
            messages.success(request, f'Imagen del carrusel creada exitosamente.')
            return redirect('gestionar_carrusel')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = ImagenCarruselForm()
    
    return render(request, 'core/crear_imagen_carrusel.html', {'form': form})


@login_required
def editar_imagen_carrusel_view(request, imagen_id):
    """Vista para editar una imagen del carrusel"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    imagen = get_object_or_404(ImagenCarrusel, id=imagen_id)
    
    if request.method == 'POST':
        form = ImagenCarruselForm(request.POST, request.FILES, instance=imagen)
        if form.is_valid():
            # Guardar datos anteriores para auditoría
            datos_anteriores = {
                'orden': imagen.orden,
                'activo': imagen.activo,
            }
            form.save()
            imagen.refresh_from_db()
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='carrusel_edit',
                modulo='contenido',
                descripcion=f'Imagen de carrusel modificada (orden: {imagen.orden})',
                objeto_afectado=f"Imagen Carrusel #{imagen.orden}",
                detalles={
                    'datos_anteriores': datos_anteriores,
                    'datos_nuevos': {
                        'orden': imagen.orden,
                        'activo': imagen.activo,
                    }
                }
            )
            messages.success(request, f'Imagen del carrusel actualizada exitosamente.')
            return redirect('gestionar_carrusel')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = ImagenCarruselForm(instance=imagen)
    
    return render(request, 'core/editar_imagen_carrusel.html', {'form': form, 'imagen': imagen})


@login_required
def eliminar_imagen_carrusel_view(request, imagen_id):
    """Vista para eliminar una imagen del carrusel"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    imagen = get_object_or_404(ImagenCarrusel, id=imagen_id)
    
    if request.method == 'POST':
        orden = imagen.orden
        # Registrar auditoría antes de eliminar
        from .utils import registrar_auditoria
        registrar_auditoria(
            usuario=request.user,
            accion='carrusel_delete',
            modulo='contenido',
            descripcion=f'Imagen de carrusel eliminada (orden: {orden})',
            objeto_afectado=f"Imagen Carrusel #{orden}",
            detalles={
                'orden': orden,
            }
        )
        # Eliminar la imagen si existe
        if imagen.imagen:
            imagen.imagen.delete()
        imagen.delete()
        messages.success(request, f'Imagen del carrusel eliminada exitosamente.')
        return redirect('gestionar_carrusel')
    
    return render(request, 'core/eliminar_imagen_carrusel.html', {'imagen': imagen})


# ==================== GESTIÓN DE EVENTOS ====================

@login_required
def gestionar_eventos_view(request):
    """Vista para gestionar eventos"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    eventos = Evento.objects.all().order_by('-fecha_evento', '-fecha_creacion')
    return render(request, 'core/gestionar_eventos.html', {'eventos': eventos})


@login_required
def crear_evento_view(request):
    """Vista para crear un nuevo evento"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    if request.method == 'POST':
        form = EventoForm(request.POST, request.FILES)
        if form.is_valid():
            evento = form.save()
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='evento_create',
                modulo='contenido',
                descripcion=f'Evento "{evento.titulo}" creado',
                objeto_afectado=evento.titulo,
                detalles={
                    'fecha_evento': str(evento.fecha_evento),
                    'activo': evento.activo,
                }
            )
            messages.success(request, f'Evento "{evento.titulo}" creado exitosamente.')
            return redirect('gestionar_eventos')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = EventoForm()
    
    return render(request, 'core/crear_evento.html', {'form': form})


@login_required
def editar_evento_view(request, evento_id):
    """Vista para editar un evento"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    evento = get_object_or_404(Evento, id=evento_id)
    
    if request.method == 'POST':
        form = EventoForm(request.POST, request.FILES, instance=evento)
        if form.is_valid():
            # Guardar datos anteriores para auditoría
            datos_anteriores = {
                'titulo': evento.titulo,
                'fecha_evento': str(evento.fecha_evento),
                'activo': evento.activo,
            }
            
            # Manejar eliminación/reemplazo de imagen
            eliminar_imagen = form.cleaned_data.get('eliminar_imagen', False)
            nueva_imagen = form.cleaned_data.get('imagen')
            
            # Guardar una referencia a la imagen actual ANTES de cualquier modificación
            # Necesitamos obtener el objeto completo antes de que se modifique
            evento_temp = Evento.objects.get(id=evento.id)
            imagen_anterior_path = None
            if evento_temp.imagen:
                imagen_anterior_path = str(evento_temp.imagen.name)
            
            # Si se marca eliminar imagen, hacerlo ANTES de guardar el formulario
            if eliminar_imagen:
                # Eliminar el archivo físico del disco si existe
                if imagen_anterior_path:
                    try:
                        # Eliminar el archivo usando el storage de Django
                        if default_storage.exists(imagen_anterior_path):
                            default_storage.delete(imagen_anterior_path)
                    except Exception as e:
                        # Si falla la eliminación, continuar de todas formas
                        logger.warning(f"Error eliminando imagen de evento: {e}")
                
                # Guardar el formulario sin commit para establecer imagen como None
                evento = form.save(commit=False)
                evento.imagen = None
                evento.save()
                
                # Forzar la actualización en la base de datos para asegurar que el campo esté vacío
                Evento.objects.filter(id=evento.id).update(imagen=None)
                evento.refresh_from_db()
                messages.success(request, f'Evento "{evento.titulo}" actualizado exitosamente. La imagen ha sido eliminada.')
            else:
                # Si no se marca eliminar, guardar normalmente
                evento = form.save()
                evento.refresh_from_db()
                messages.success(request, f'Evento "{evento.titulo}" actualizado exitosamente.')
            
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='evento_edit',
                modulo='contenido',
                descripcion=f'Evento "{evento.titulo}" modificado',
                objeto_afectado=evento.titulo,
                detalles={
                    'datos_anteriores': datos_anteriores,
                    'datos_nuevos': {
                        'titulo': evento.titulo,
                        'fecha_evento': str(evento.fecha_evento),
                        'activo': evento.activo,
                    }
                }
            )
            return redirect('gestionar_eventos')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = EventoForm(instance=evento)
    
    return render(request, 'core/editar_evento.html', {'form': form, 'evento': evento})


@login_required
def eliminar_evento_view(request, evento_id):
    """Vista para eliminar un evento"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    evento = get_object_or_404(Evento, id=evento_id)
    
    if request.method == 'POST':
        titulo = evento.titulo
        fecha_evento = str(evento.fecha_evento)
        # Registrar auditoría antes de eliminar
        from .utils import registrar_auditoria
        registrar_auditoria(
            usuario=request.user,
            accion='evento_delete',
            modulo='contenido',
            descripcion=f'Evento "{titulo}" eliminado',
            objeto_afectado=titulo,
            detalles={
                'fecha_evento': fecha_evento,
            }
        )
        # Eliminar la imagen si existe
        if evento.imagen:
            evento.imagen.delete()
        evento.delete()
        messages.success(request, f'Evento "{titulo}" eliminado exitosamente.')
        return redirect('gestionar_eventos')
    
    return render(request, 'core/eliminar_evento.html', {'evento': evento})


# ==================== GESTIÓN DE NOTICIAS ====================

@login_required
def gestionar_noticias_view(request):
    """Vista para gestionar noticias"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    noticias = Noticia.objects.all().order_by('-fecha_publicacion', '-fecha_creacion')
    return render(request, 'core/gestionar_noticias.html', {'noticias': noticias})


@login_required
def crear_noticia_view(request):
    """Vista para crear una nueva noticia"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    if request.method == 'POST':
        form = NoticiaForm(request.POST, request.FILES)
        if form.is_valid():
            noticia = form.save()
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='noticia_create',
                modulo='contenido',
                descripcion=f'Noticia "{noticia.titulo}" creada',
                objeto_afectado=noticia.titulo,
                detalles={
                    'fecha_publicacion': str(noticia.fecha_publicacion),
                    'activo': noticia.activo,
                }
            )
            messages.success(request, f'Noticia "{noticia.titulo}" creada exitosamente.')
            return redirect('gestionar_noticias')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = NoticiaForm()
    
    return render(request, 'core/crear_noticia.html', {'form': form})


@login_required
def editar_noticia_view(request, noticia_id):
    """Vista para editar una noticia"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    noticia = get_object_or_404(Noticia, id=noticia_id)
    
    if request.method == 'POST':
        form = NoticiaForm(request.POST, request.FILES, instance=noticia)
        if form.is_valid():
            # Guardar datos anteriores para auditoría
            datos_anteriores = {
                'titulo': noticia.titulo,
                'fecha_publicacion': str(noticia.fecha_publicacion),
                'activo': noticia.activo,
            }
            
            # Manejar eliminación/reemplazo de imagen
            eliminar_imagen = form.cleaned_data.get('eliminar_imagen', False)
            nueva_imagen = form.cleaned_data.get('imagen')
            
            # Guardar una referencia a la imagen actual ANTES de cualquier modificación
            # Necesitamos obtener el objeto completo antes de que se modifique
            noticia_temp = Noticia.objects.get(id=noticia.id)
            imagen_anterior_path = None
            if noticia_temp.imagen:
                imagen_anterior_path = str(noticia_temp.imagen.name)
            
            # Si se marca eliminar imagen, hacerlo ANTES de guardar el formulario
            if eliminar_imagen:
                # Eliminar el archivo físico del disco si existe
                if imagen_anterior_path:
                    try:
                        # Eliminar el archivo usando el storage de Django
                        if default_storage.exists(imagen_anterior_path):
                            default_storage.delete(imagen_anterior_path)
                    except Exception as e:
                        # Si falla la eliminación, continuar de todas formas
                        logger.warning(f"Error eliminando imagen de evento: {e}")
                
                # Guardar el formulario sin commit para establecer imagen como None
                noticia = form.save(commit=False)
                noticia.imagen = None
                noticia.save()
                
                # Forzar la actualización en la base de datos para asegurar que el campo esté vacío
                Noticia.objects.filter(id=noticia.id).update(imagen=None)
                noticia.refresh_from_db()
                messages.success(request, f'Noticia "{noticia.titulo}" actualizada exitosamente. La imagen ha sido eliminada.')
            else:
                # Si no se marca eliminar, guardar normalmente
                noticia = form.save()
                noticia.refresh_from_db()
                messages.success(request, f'Noticia "{noticia.titulo}" actualizada exitosamente.')
            
            # Registrar auditoría
            from .utils import registrar_auditoria
            registrar_auditoria(
                usuario=request.user,
                accion='noticia_edit',
                modulo='contenido',
                descripcion=f'Noticia "{noticia.titulo}" modificada',
                objeto_afectado=noticia.titulo,
                detalles={
                    'datos_anteriores': datos_anteriores,
                    'datos_nuevos': {
                        'titulo': noticia.titulo,
                        'fecha_publicacion': str(noticia.fecha_publicacion),
                        'activo': noticia.activo,
                    }
                }
            )
            return redirect('gestionar_noticias')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = NoticiaForm(instance=noticia)
    
    return render(request, 'core/editar_noticia.html', {'form': form, 'noticia': noticia})


@login_required
def eliminar_noticia_view(request, noticia_id):
    """Vista para eliminar una noticia"""
    if request.user.es_colaborador and request.user.cambio_password_requerido:
        return redirect('cambiar_password')
    if not (request.user._es_administrador or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para acceder a esta sección')
        return redirect('panel')
    
    noticia = get_object_or_404(Noticia, id=noticia_id)
    
    if request.method == 'POST':
        titulo = noticia.titulo
        fecha_publicacion = str(noticia.fecha_publicacion)
        # Registrar auditoría antes de eliminar
        from .utils import registrar_auditoria
        registrar_auditoria(
            usuario=request.user,
            accion='noticia_delete',
            modulo='contenido',
            descripcion=f'Noticia "{titulo}" eliminada',
            objeto_afectado=titulo,
            detalles={
                'fecha_publicacion': fecha_publicacion,
            }
        )
        noticia.delete()
        messages.success(request, f'Noticia "{titulo}" eliminada exitosamente.')
        return redirect('gestionar_noticias')
    
    return render(request, 'core/eliminar_noticia.html', {'noticia': noticia})
